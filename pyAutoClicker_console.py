import logging
import pandas as pd
import pyautogui

from openpyxl import load_workbook
from pynput.mouse import Listener as MouseListener
from pynput.keyboard import Listener as KeyboardListener
from pynput.keyboard import KeyCode
from time import sleep as pause


def mouse_monitor():

# filemode - w - nadpisuje plik od początku

    logging.basicConfig(filename="mysz.txt",filemode="w" ,level=logging.DEBUG, format="%(message)s")
    stop_key = KeyCode(char="s")
    def on_click(x,y, button, pressed):
        if pressed:
            logging.info("klik \t{0}\t {1}\t{2}".format(x, y, button))

# akcja klawisza stopu

    def on_press(key):
        if key == stop_key:
            listener.stop()

# monitoring myszy i klawiatury jednoczesnie

    with MouseListener(on_click=on_click) as listener:
        with KeyboardListener(on_press=on_press) as listener:
            listener.join()


def convert_log_to_xls():
    log = pd.read_table("mysz.txt")
    log.to_excel("mysz.xlsx", index=False, header=True)

#dodanie wiersza z tytułem umozliwia uzycie funkcji loc biblioteki pandas, ktora traktuje pierwszy wiersz jako tytulowy

def add_row_title():
    skoroszyt = load_workbook("mysz.xlsx")
    arkusz = skoroszyt.active
    arkusz.insert_rows(1)
    arkusz.cell(1,1,"KolumnaA")
    arkusz.cell(1,2,"KolumnaB")
    arkusz.cell(1,3,"KolumnaC")
    arkusz.cell(1,4,"KolumnaD")
    skoroszyt.save("mysz.xlsx")

def read_x():
    x = pd.read_excel("mysz.xlsx")
    licznik = 0
    ilosc_wierszy = x["KolumnaB"].size
    tablicax = []
    while licznik < ilosc_wierszy:
        asd = x["KolumnaB"].loc[licznik]
        licznik += 1
        tablicax.append(asd)
    return tablicax

def read_y():
    y = pd.read_excel("mysz.xlsx")
    licznik = 0
    ilosc_wierszy = y["KolumnaC"].size
    tablicay = []
    while licznik < ilosc_wierszy:
        asd = y["KolumnaC"].loc[licznik]
        licznik += 1
        tablicay.append(asd)
    return tablicay

def read_button():
    od = pd.read_excel("mysz.xlsx")
    licznik = 0
    ilosc_wierszy = od["KolumnaD"].size
    tablicaod = []
    while licznik < ilosc_wierszy:
        asd = od["KolumnaD"].loc[licznik]
        licznik += 1
        string_przycisku = str(asd)
        string_przycisku = string_przycisku.replace("Button.left", "left")
        string_przycisku = string_przycisku.replace("Button.right", "right")
        tablicaod.append(string_przycisku)
    return tablicaod

# funkcja korzysta z odczytx i odczyty

def start_bota():
    ile_razy_wykonac = int(input("Ile petli wykonac:"))
    convert_log_to_xls()
    add_row_title()
    licznikx = 0
    liczniky = 0
    licznikod = 0
    licznik_wykonania = pd.read_excel("mysz.xlsx")["KolumnaB"].size-1
    for x in range(ile_razy_wykonac):
        while licznik_wykonania >= liczniky:
            y = (read_y()[liczniky])
            x = (read_x()[licznikx])
            button = (read_button()[licznikod])
            pyautogui.click(x,y, button=button, interval=2)
            licznikx += 1
            liczniky += 1
            #pause(2)

def menu():
    print('''
    Witaj w programie pyAutoClicker, który umożliwi Ci odtworzenie Twoich kliknięć 
    Program pobiera informacje o użytym przycisku oraz współrzędne miejsca kliknięcia
    Aby zakończyć rejestrowanie nacisnij klawisz 's'   
    ''')
    print('''
    1. Rejestrowanie zdarzen
    
    2. Odtwarzanie zdarzen
    
    3. Zakoncz program
    ''')



while True:
    menu()
    choice = input("Wybierz tryb pracy:")
    if choice == "1":
        mouse_monitor()
        continue
    elif choice == "2":
        start_bota()
    else:
        print("Dziękuję za skorzystanie z programu.")
        break

