from pynput.mouse import Listener as MouseListener
from pynput.keyboard import Listener as KeyboardListener
from pynput.keyboard import KeyCode
from openpyxl import load_workbook
from time import sleep as pause
from tkinter import *
import logging
import pandas as pd
import pyautogui

def mouse_monitor():

# filemode - w - nadpisuje plik od początku

    logging.basicConfig(filename="mysz.txt",filemode="w" ,level=logging.DEBUG, format="%(message)s")
    stop_key = KeyCode(char="s")
    def on_click(x,y, button, pressed):
        if pressed:
            logging.info("klik \t{0}\t {1}\t{2}".format(x, y, button))

# akcja klawisza stopu

    def on_press(key):
        if key == stop_key:
            listener.stop()

# monitoring myszy i klawiatury jednoczesnie

    with MouseListener(on_click=on_click) as listener:
        with KeyboardListener(on_press=on_press) as listener:
            listener.join()

def convert_log_to_xls():
    log = pd.read_table("mysz.txt")
    log.to_excel("mysz.xlsx", index=False, header=True)

#dodanie wiersza z tytułem umozliwia uzycie funkcji loc biblioteki pandas, ktora traktuje pierwszy wiersz jako tytulowy

def add_row_title():
    skoroszyt = load_workbook("mysz.xlsx")
    arkusz = skoroszyt.active
    arkusz.insert_rows(1)
    arkusz.cell(1,1,"KolumnaA")
    arkusz.cell(1,2,"KolumnaB")
    arkusz.cell(1,3,"KolumnaC")
    arkusz.cell(1,4,"KolumnaD")
    skoroszyt.save("mysz.xlsx")

def read_x():
    x = pd.read_excel("mysz.xlsx")
    licznik = 0
    ilosc_wierszy = x["KolumnaB"].size
    tablicax = []
    while licznik < ilosc_wierszy:
        asd = x["KolumnaB"].loc[licznik]
        licznik += 1
        tablicax.append(asd)
    return tablicax

def read_y():
    y = pd.read_excel("mysz.xlsx")
    licznik = 0
    ilosc_wierszy = y["KolumnaC"].size
    tablicay = []
    while licznik < ilosc_wierszy:
        asd = y["KolumnaC"].loc[licznik]
        licznik += 1
        tablicay.append(asd)
    return tablicay

def read_button():
    od = pd.read_excel("mysz.xlsx")
    licznik = 0
    ilosc_wierszy = od["KolumnaD"].size
    tablicaod = []
    while licznik < ilosc_wierszy:
        asd = od["KolumnaD"].loc[licznik]
        licznik += 1
        string_przycisku = str(asd)
        string_przycisku = string_przycisku.replace("Button.left", "left")
        string_przycisku = string_przycisku.replace("Button.right", "right")
        tablicaod.append(string_przycisku)
    return tablicaod

# funkcja korzysta z odczytx i odczyty

def start_bota():
    convert_log_to_xls()
    add_row_title()
    licznikx = 0
    liczniky = 0
    licznikod = 0
    licznik_wykonania = pd.read_excel("mysz.xlsx")["KolumnaB"].size-1
    while licznik_wykonania > ile_razy_wykonac:
        while licznik_wykonania >= liczniky:
            y = (read_y()[liczniky])
            x = (read_x()[licznikx])
            button = (read_button()[licznikod])
            pyautogui.click(x,y, button=button, interval=2)
            licznikx += 1
            liczniky += 1
            #pause(2)

def main():

    window = Tk()
    window.title("Rejestrator bota")

    opcja1 = Button(window, text="Rejestrator bota", activeforeground= "#FF0000", command=mouse_monitor).grid(sticky=E+W, padx = 5, pady = 5)
    opcja2 = Button(window,text="Uruchom bota", activeforeground= "#FF0000", command=start_bota).grid(sticky=E+W, padx = 5, pady = 5)
    tytul = Label(window, text="Wybierz ile razy zostanie powtórzone zadanie:").grid(sticky=E+W, padx=5, pady=5)
    ilosc_powtorzen = Spinbox(window, from_=1, to=20).grid(sticky=E + W, padx=5, pady=5)
    opcja3 = Button(window, text="Zakończ", command=window.quit).grid(sticky=E+W, padx = 5, pady = 5)
    ilosc_powtorzen = IntVar()
    ile_razy_wykonac = ilosc_powtorzen.get()

    window.mainloop()

if __name__ == "__main__":
    main()